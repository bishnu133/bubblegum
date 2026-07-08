"""Generate examples/sample_scenarios_banking.xlsx — a second, different-domain
sample for trying `bubblegum convert` (online banking, nothing to do with any
particular project). Filenames in the output derive purely from the Feature/Epic
column, proving the generator is data-driven.

Run:  python examples/make_sample_banking.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["#", "Feature/Epic", "Test Scenario", "User Persona",
           "Functional Jira Story", "Verify", "Result", "Remarks"]

ROWS = [
    (1, "[F][Web] Account overview", "Verify account balance is displayed after login",
     "Retail customer", "BANK-101",
     'Given I am logged in as a "Retail customer"\n'
     "And I open the Accounts page\n"
     "When I select the Savings account\n"
     "Then I see the Available balance"),
    (2, "[F][Web] Fund transfer", "Verify a transfer between own accounts succeeds",
     "Retail customer", "BANK-201",
     'Given I am logged in as a "Retail customer"\n'
     "And I open the Transfer page\n"
     'When I select "Savings" from the From account dropdown\n'
     'And I select "Current" from the To account dropdown\n'
     'And I enter "150.00" into the Amount field\n'
     "And I click the Continue button\n"
     "And I click the Confirm button\n"
     "Then I see the Transfer successful message"),
    (3, "[F][Web] Fund transfer", "Verify transfer is rejected when amount exceeds balance",
     "Retail customer", "BANK-202",
     "Given I open the Transfer page\n"
     'When I enter "999999" into the Amount field\n'
     "And I click the Continue button\n"
     "Then I see the Insufficient funds error"),
    (4, "[F][Web] Bill payment", "Verify a registered biller can be paid",
     "Retail customer", "BANK-301",
     "Given I open the Bill payment page\n"
     'When I select "Electricity Board" from the Biller dropdown\n'
     'And I enter "80.50" into the Amount field\n'
     "And I click the Pay button\n"
     "Then I see the Payment scheduled confirmation"),
    (5, "[F][Backend] Interest accrual", "Verify monthly interest is credited to savings",
     "Retail customer", "BANK-400",
     "Given a savings account with a 100000 balance\n"
     "When the monthly interest job runs\n"
     "Then the account is credited with the accrued interest"),
]


def build() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UAT Scenario"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(list(r) + ["", ""])

    header_fill = PatternFill("solid", fgColor="2E7D32")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    widths = {"A": 5, "B": 30, "C": 50, "D": 18, "E": 16, "F": 72, "G": 10, "H": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=cell.column_letter == "F", vertical="top")
    ws.freeze_panes = "A2"

    out = Path(__file__).with_name("sample_scenarios_banking.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    print("wrote", build())
