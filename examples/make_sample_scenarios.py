"""Generate examples/sample_scenarios_template.xlsx for trying `bubblegum convert`.

Run:  python examples/make_sample_scenarios.py

Synthetic data (no proprietary content). The scenarios are written in the
concrete-Gherkin style from docs/authoring-scenarios-style-guide.md, plus one
[Backend] row (flagged non-UI) and one deliberately abstract row (to show the
honest NEEDS_DATA / MANUAL markers). Column layout mirrors a real UAT sheet.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["#", "Feature/Epic", "Test Scenario", "User Persona",
           "Functional Jira Story", "Verify", "Result", "Remarks"]

ROWS = [
    (1, "[F][Web] Login", "Verify login with valid credentials succeeds",
     "Registered user", "DEMO-101",
     "Given I open the Login page\n"
     'When I enter "tomsmith" into the Username field\n'
     'And I enter "SuperSecretPassword!" into the Password field\n'
     "And I click the Login button\n"
     "Then I see the Secure Area heading"),
    (2, "[F][Web] Login", "Verify login with an invalid password is rejected",
     "Registered user", "DEMO-101",
     "Given I open the Login page\n"
     'When I enter "tomsmith" into the Username field\n'
     'And I enter "wrongpass" into the Password field\n'
     "And I click the Login button\n"
     "Then I see the Your password is invalid error"),
    (3, "[F][Web] Checkout", "Verify a valid coupon applies a discount",
     "Shopper", "DEMO-201",
     'Given I am logged in as a "Shopper"\n'
     "And I open the Checkout page\n"
     'When I enter "SAVE10" into the Coupon code field\n'
     "And I click the Apply button\n"
     "Then I see the Discount applied message"),
    (4, "[F][Web] Checkout", "Verify removing an item updates the cart total",
     "Shopper", "DEMO-202",
     "Given I open the Cart page\n"
     "And the cart has 2 items\n"
     "When I click the Remove button for the first item\n"
     "Then I see the Cart total updated"),
    (5, "[F][Web] Profile", "Verify a user can update their display name",
     "Registered user", "DEMO-301",
     'Given I am logged in as a "Registered user"\n'
     "And I open the Profile page\n"
     'When I enter "Jane Doe" into the Display name field\n'
     "And I click the Save button\n"
     "Then I see the Profile updated confirmation"),
    (6, "[F][Backend] Rewards engine", "Verify points accrue after a purchase",
     "Registered user", "DEMO-400",
     "Given a user with 0 reward points\n"
     "When an order of 100 dollars is completed\n"
     "Then the user has 100 reward points"),
    (7, "[F][Web] Badge album", "Verify the badge album shows earned badges",
     "H365 user", "DEMO-500",
     "Given a H365 user viewing their badge album page\n"
     "When there is at least 1 earned badge\n"
     "Then the user will be able to see their badges correctly"),
]


def build() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UAT Scenario"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(list(r) + ["", ""])

    header_fill = PatternFill("solid", fgColor="4F81BD")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    widths = {"A": 5, "B": 34, "C": 46, "D": 18, "E": 16, "F": 70, "G": 10, "H": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            wrap = cell.column_letter == "F"
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    ws.freeze_panes = "A2"

    out = Path(__file__).with_name("sample_scenarios_template.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    print("wrote", build())
