"""
bubblegum/convert/ingest.py
===========================
Read a spreadsheet of manual test scenarios into RawScenario objects.

Column mapping is driven entirely by the ConvertProfile so a team can use any
header names. We match headers case-insensitively and tolerate surrounding
whitespace. ``openpyxl`` is an optional dependency (the ``convert`` extra); we
import it lazily and raise a clear, actionable error if it is missing.
"""

from __future__ import annotations

from pathlib import Path

from bubblegum.convert.models import RawScenario
from bubblegum.convert.profile import ConvertProfile


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:  # pragma: no cover - import-guard
        raise ImportError(
            "Reading .xlsx files needs openpyxl. Install the convert extra:\n"
            '    pip install "bubblegum-ai[convert]"\n'
            "or install openpyxl directly: pip install openpyxl"
        ) from exc
    import openpyxl

    return openpyxl


def _header_index(header_cells: list, columns: dict[str, str]) -> dict[str, int]:
    """Map logical field name -> 0-based column index using header text.

    ``columns`` is {logical_name: header_label}. Matching is case-insensitive
    and whitespace-insensitive. Missing headers are simply absent from the map.
    """
    normalized = {
        str(cell).strip().casefold(): idx
        for idx, cell in enumerate(header_cells)
        if cell is not None and str(cell).strip()
    }
    resolved: dict[str, int] = {}
    for logical, label in columns.items():
        key = str(label).strip().casefold()
        if key in normalized:
            resolved[logical] = normalized[key]
    return resolved


def read_workbook(
    path: str | Path,
    profile: ConvertProfile | None = None,
) -> list[RawScenario]:
    """Read scenarios from an .xlsx workbook into RawScenario objects.

    Sheet selection: ``input.sheet`` (single) or ``input.sheets`` (list) restrict
    which sheets are read; otherwise every sheet that has the steps column is
    processed. Each RawScenario is tagged with its sheet. Rows with an empty
    steps cell are skipped (usually spacer rows).
    """
    profile = profile or ConvertProfile()
    openpyxl = _require_openpyxl()

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    wanted = _wanted_sheets(profile)
    if wanted is not None:
        missing = [s for s in wanted if s not in wb.sheetnames]
        if missing:
            raise ValueError(
                f"Sheet(s) {missing} not found. Available: {wb.sheetnames}"
            )
        worksheets = [wb[s] for s in wanted]
    else:
        worksheets = list(wb.worksheets)

    header_row = profile.input.header_row
    scenarios: list[RawScenario] = []
    matched_any = False

    for ws in worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < header_row:
            continue
        header_cells = list(rows[header_row - 1])
        index = _header_index(header_cells, profile.input.columns)
        steps_col = index.get("steps")
        if steps_col is None:
            # When a specific sheet was requested it must have the column;
            # when scanning all sheets, silently skip non-scenario sheets.
            if wanted is not None:
                raise ValueError(
                    f"Steps column {profile.input.columns.get('steps')!r} not found in "
                    f"sheet {ws.title!r} headers: {[c for c in header_cells if c]}. "
                    "Set convert.input.columns.steps in bubblegum.convert.yaml."
                )
            continue
        matched_any = True

        for offset, row in enumerate(rows[header_row:], start=header_row + 1):
            steps_val = row[steps_col] if steps_col < len(row) else None
            if steps_val is None or not str(steps_val).strip():
                continue
            fields: dict[str, str] = {}
            for logical, col in index.items():
                if logical == "steps":
                    continue
                val = row[col] if col < len(row) else None
                if val is not None:
                    fields[logical] = str(val).strip()
            scenarios.append(
                RawScenario(row=offset, steps_text=str(steps_val), fields=fields, sheet=ws.title)
            )

    wb.close()

    if not matched_any and wanted is None:
        raise ValueError(
            f"Steps column {profile.input.columns.get('steps')!r} not found in any sheet. "
            "Set convert.input.columns.steps in bubblegum.convert.yaml."
        )
    return scenarios


def _wanted_sheets(profile: ConvertProfile) -> list[str] | None:
    """Explicit sheet allow-list, or None to scan all sheets."""
    if profile.input.sheets:
        return list(profile.input.sheets)
    if profile.input.sheet:
        return [profile.input.sheet]
    return None
